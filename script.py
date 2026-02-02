import argparse
import os
import re
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook
import markdown

# helpers

def slugify(text):
    """Create a safe id for HTML anchors."""
    text = text.lower();
    text = re.sub(r"[^\w\s-]", "", text); # regex is AI-generated
    text = re.sub(r"\s+", "-", text.strip()); # regex is AI-generated
    return text;


def read_markdown(path):
    """Convert Markdown file to HTML."""
    with open(path, "r", encoding="utf-8") as f:
        return markdown.markdown(f.read());

# general_dmp.xlsx

def parse_general_dmp(path):
    """
    Parses general information about DMP from general_dmp.xlsx.
    Value of the 2st cell in the row starting with "Title of DMP" is translated into the title of the document.
    Values from column A are translated to <h3> headings.
    Values from the next columns in the same row are translated into <p> paragraphs.
    """
    wb = load_workbook(path);
    ws = wb.active;

    title = None;
    html_parts = [];

    for row in ws.iter_rows(min_row=2, values_only=True):
        heading = row[0];
        if not heading:
            continue;

        if heading == "Title of DMP":
            title = row[1];
            continue;

        html_parts.append(f"<h3>{heading}</h3>");

        for cell in row[1:]:
            if cell:
                html_parts.append(f"<p>{cell}</p>");

    return title, "\n".join(html_parts);



# data_table.xlsx

def parse_data_table(path):
    """
    Parses DMP data for the research output from data_table.xlsx.
    Headers are taken from cells in the first row.
    Paragraphs for sections are taken from the cells in the same column.
    Paragraph begins with the name of the research output specified in the same row in column A.
    """
    wb = load_workbook(path, data_only=True);
    ws = wb.active;

    # parse heading definitions from first row
    heading_defs = {};
    heading_levels = {};

    for col in range(2, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value;
        if not cell_value:
            continue;

        matches = re.match(r"((\d+\.)+)\s*(.*)", cell_value); # regex is AI-generated
        if not matches:
            continue;

        numbering = matches.group(1);
        text = matches.group(3);
        level = numbering.count(".");

        heading_defs[col] = text;
        heading_levels[col] = level;

    html_parts = [];

    for row in range(2, ws.max_row + 1):
        output_name = ws.cell(row=row, column=1).value;
        if not output_name:
            continue;

        section_id = slugify(output_name);
        html_parts.append(f'<h1 id="{section_id}">{output_name}</h1>');

        collected = defaultdict(list);

        for col, heading in heading_defs.items():
            value = ws.cell(row=row, column=col).value;
            level = heading_levels[col];

            if level == 1:
                collected["h1"].append((heading, value));
            elif level == 2:
                collected["h2"].append((heading, value));
            elif level == 3:
                collected["h3"].append((heading, value));

        # navigation list (links to h1)
        nav_items = [];

        for heading, _ in collected["h1"]:
            hid = slugify(f"{output_name}-{heading}");
            nav_items.append(f'<li><a href="#{hid}">{heading}</a></li>');

        if nav_items:
            html_parts.append("<ul>");
            html_parts.extend(nav_items);
            html_parts.append("</ul>");

        # Render sections
        for col, heading in heading_defs.items():
            value = ws.cell(row=row, column=col).value;
            level = heading_levels[col];

            if level == 3 and value == "N/A":
                continue;

            if level == 2:
                has_children = any(
                    v != "N/A"
                    for h, v in collected["h3"]
                    if h.startswith(heading)
                );
                if not has_children:
                    continue;

            hid = slugify(f"{output_name}-{heading}");
            html_parts.append(f'<h{level} id="{hid}">{heading}</h{level}>');

            if level == 3 and value:
                html_parts.append(f"<p>{value}</p>");

    return "\n".join(html_parts);



# main


def main():
    parser = argparse.ArgumentParser();
    parser.add_argument("--data", help="Data subfolder", default=".");
    args = parser.parse_args();

    base = args.data;

    title, general_html = parse_general_dmp(
        os.path.join(base, "general_dmp.xlsx")
    );

    html = [
        "<!DOCTYPE html>",
        "<html>",
        "<head>",
        '<meta charset="utf-8">',
        f"<title>{title or 'DMP'}</title>",
        '<link rel="stylesheet" href="style.css">',
        "</head>",
        "<body>",
    ];

    html.append(general_html);

    for md in ["introduction.md", "data.md"]:
        html.append(read_markdown(os.path.join(base, md)));

    html.append(parse_data_table(os.path.join(base, "data_table.xlsx")));

    for md in ["conclusion.md", "acknowledgements.md", "references.md"]:
        html.append(read_markdown(os.path.join(base, md)));

    html.extend(["</body>", "</html>"]);

    with open("output/dmp.html", "w", encoding="utf-8") as f:
        f.write("\n".join(html));


if __name__ == "__main__":
    main();
